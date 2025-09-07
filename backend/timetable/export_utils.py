"""
Timetable export utilities for various formats
"""

import io
import json
from datetime import datetime, timedelta
from typing import Dict, List, Optional
from django.http import HttpResponse
from django.template.loader import render_to_string
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
from .models import Timetable, TimetableSession


class TimetableExporter:
    """
    Utility class for exporting timetables in various formats
    """
    
    def __init__(self, timetable: Timetable):
        self.timetable = timetable
        self.sessions = timetable.sessions.select_related(
            'subject', 'teacher__user', 'room', 'class_group'
        ).order_by('day_of_week', 'start_time')
    
    def export_pdf(self, view_type: str = 'general') -> HttpResponse:
        """
        Export timetable as PDF
        """
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        title = f"{self.timetable.institution.name} - {self.timetable.name}"
        story.append(Paragraph(title, title_style))
        
        # Subtitle
        subtitle = f"Academic Year: {self.timetable.academic_year} | Semester: {self.timetable.semester}"
        story.append(Paragraph(subtitle, styles['Normal']))
        story.append(Spacer(1, 20))
        
        if view_type == 'general':
            self._add_general_timetable_to_pdf(story, styles)
        elif view_type == 'teacher':
            self._add_teacher_wise_timetable_to_pdf(story, styles)
        elif view_type == 'class':
            self._add_class_wise_timetable_to_pdf(story, styles)
        
        doc.build(story)
        buffer.seek(0)
        
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{self.timetable.name}.pdf"'
        return response
    
    def _add_general_timetable_to_pdf(self, story, styles):
        """Add general timetable view to PDF"""
        # Create time slot grid
        time_slots = self._get_time_slots()
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        working_days = self.timetable.institution.working_days or [0, 1, 2, 3, 4]
        
        # Create table data
        table_data = []
        header_row = ['Time'] + [days[day] for day in working_days]
        table_data.append(header_row)
        
        for time_slot in time_slots:
            row = [f"{time_slot['start_time']} - {time_slot['end_time']}"]
            
            for day in working_days:
                sessions_at_slot = self.sessions.filter(
                    day_of_week=day,
                    start_time=time_slot['start_time']
                )
                
                if sessions_at_slot.exists():
                    session_texts = []
                    for session in sessions_at_slot:
                        text = f"{session.class_group}\n{session.subject.code}"
                        if session.teacher:
                            text += f"\n{session.teacher.user.first_name[0]}.{session.teacher.user.last_name}"
                        if session.room:
                            text += f"\n{session.room.code}"
                        session_texts.append(text)
                    row.append('\n---\n'.join(session_texts))
                else:
                    row.append('')
            
            table_data.append(row)
        
        # Create and style table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
    
    def _add_teacher_wise_timetable_to_pdf(self, story, styles):
        """Add teacher-wise timetable to PDF"""
        teachers = set(session.teacher for session in self.sessions if session.teacher)
        
        for teacher in teachers:
            teacher_sessions = self.sessions.filter(teacher=teacher)
            
            # Teacher header
            teacher_title = f"Teacher: {teacher.user.get_full_name()} ({teacher.employee_id})"
            story.append(Paragraph(teacher_title, styles['Heading2']))
            story.append(Spacer(1, 10))
            
            # Create teacher's timetable
            self._create_individual_timetable_table(story, teacher_sessions, 'teacher')
            story.append(Spacer(1, 20))
    
    def _add_class_wise_timetable_to_pdf(self, story, styles):
        """Add class-wise timetable to PDF"""
        classes = set(session.class_group for session in self.sessions)
        
        for class_group in classes:
            class_sessions = self.sessions.filter(class_group=class_group)
            
            # Class header
            class_title = f"Class: {class_group}"
            story.append(Paragraph(class_title, styles['Heading2']))
            story.append(Spacer(1, 10))
            
            # Create class timetable
            self._create_individual_timetable_table(story, class_sessions, 'class')
            story.append(Spacer(1, 20))
    
    def _create_individual_timetable_table(self, story, sessions, view_type):
        """Create individual timetable table for teacher or class"""
        time_slots = self._get_time_slots()
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        working_days = self.timetable.institution.working_days or [0, 1, 2, 3, 4]
        
        table_data = []
        header_row = ['Time'] + [days[day] for day in working_days]
        table_data.append(header_row)
        
        for time_slot in time_slots:
            row = [f"{time_slot['start_time']} - {time_slot['end_time']}"]
            
            for day in working_days:
                session = sessions.filter(
                    day_of_week=day,
                    start_time=time_slot['start_time']
                ).first()
                
                if session:
                    if view_type == 'teacher':
                        text = f"{session.class_group}\n{session.subject.code}"
                        if session.room:
                            text += f"\n{session.room.code}"
                    else:  # class view
                        text = f"{session.subject.code}"
                        if session.teacher:
                            text += f"\n{session.teacher.user.first_name[0]}.{session.teacher.user.last_name}"
                        if session.room:
                            text += f"\n{session.room.code}"
                    row.append(text)
                else:
                    row.append('')
            
            table_data.append(row)
        
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(table)
    
    def export_excel(self, view_type: str = 'general') -> HttpResponse:
        """
        Export timetable as Excel file
        """
        workbook = openpyxl.Workbook()
        
        if view_type == 'general':
            self._create_general_excel_sheet(workbook)
        elif view_type == 'teacher':
            self._create_teacher_wise_excel_sheets(workbook)
        elif view_type == 'class':
            self._create_class_wise_excel_sheets(workbook)
        
        # Remove default sheet if we created others
        if len(workbook.worksheets) > 1:
            workbook.remove(workbook.worksheets[0])
        
        # Save to buffer
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self.timetable.name}.xlsx"'
        return response
    
    def _create_general_excel_sheet(self, workbook):
        """Create general timetable Excel sheet"""
        ws = workbook.active
        ws.title = "General Timetable"
        
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{self.timetable.institution.name} - {self.timetable.name}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        time_slots = self._get_time_slots()
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        working_days = self.timetable.institution.working_days or [0, 1, 2, 3, 4]
        
        # Create headers
        headers = ['Time'] + [days[day] for day in working_days]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Fill data
        for row_idx, time_slot in enumerate(time_slots, 4):
            ws.cell(row=row_idx, column=1, value=f"{time_slot['start_time']} - {time_slot['end_time']}")
            
            for col_idx, day in enumerate(working_days, 2):
                sessions_at_slot = self.sessions.filter(
                    day_of_week=day,
                    start_time=time_slot['start_time']
                )
                
                if sessions_at_slot.exists():
                    session_texts = []
                    for session in sessions_at_slot:
                        text = f"{session.class_group} - {session.subject.code}"
                        if session.teacher:
                            text += f"\n{session.teacher.user.get_full_name()}"
                        if session.room:
                            text += f"\n{session.room.code}"
                        session_texts.append(text)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value='\n---\n'.join(session_texts))
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
                    
                    # Color coding based on subject type
                    if sessions_at_slot.first().subject:
                        subject_type = sessions_at_slot.first().subject.type
                        if subject_type == 'core':
                            cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
                        elif subject_type == 'elective':
                            cell.fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
                        elif subject_type == 'lab':
                            cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_teacher_wise_excel_sheets(self, workbook):
        """Create teacher-wise Excel sheets"""
        teachers = set(session.teacher for session in self.sessions if session.teacher)
        
        for teacher in teachers:
            teacher_sessions = self.sessions.filter(teacher=teacher)
            sheet_name = f"{teacher.user.first_name} {teacher.user.last_name}"[:31]  # Excel sheet name limit
            ws = workbook.create_sheet(title=sheet_name)
            
            self._fill_individual_excel_sheet(ws, teacher_sessions, f"Teacher: {teacher.user.get_full_name()}", 'teacher')
    
    def _create_class_wise_excel_sheets(self, workbook):
        """Create class-wise Excel sheets"""
        classes = set(session.class_group for session in self.sessions)
        
        for class_group in classes:
            class_sessions = self.sessions.filter(class_group=class_group)
            sheet_name = str(class_group)[:31]  # Excel sheet name limit
            ws = workbook.create_sheet(title=sheet_name)
            
            self._fill_individual_excel_sheet(ws, class_sessions, f"Class: {class_group}", 'class')
    
    def _fill_individual_excel_sheet(self, ws, sessions, title, view_type):
        """Fill individual Excel sheet with timetable data"""
        # Title
        ws.merge_cells('A1:H1')
        ws['A1'] = title
        ws['A1'].font = Font(size=14, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        time_slots = self._get_time_slots()
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        working_days = self.timetable.institution.working_days or [0, 1, 2, 3, 4]
        
        headers = ['Time'] + [days[day] for day in working_days]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        
        # Fill data
        for row_idx, time_slot in enumerate(time_slots, 4):
            ws.cell(row=row_idx, column=1, value=f"{time_slot['start_time']} - {time_slot['end_time']}")
            
            for col_idx, day in enumerate(working_days, 2):
                session = sessions.filter(
                    day_of_week=day,
                    start_time=time_slot['start_time']
                ).first()
                
                if session:
                    if view_type == 'teacher':
                        text = f"{session.class_group}\n{session.subject.code}"
                        if session.room:
                            text += f"\n{session.room.code}"
                    else:  # class view
                        text = f"{session.subject.code}"
                        if session.teacher:
                            text += f"\n{session.teacher.user.get_full_name()}"
                        if session.room:
                            text += f"\n{session.room.code}"
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=text)
                    cell.alignment = Alignment(wrap_text=True, vertical='center')
    
    def export_ics(self) -> HttpResponse:
        """
        Export timetable as ICS calendar file
        """
        ics_content = self._generate_ics_content()
        
        response = HttpResponse(ics_content, content_type='text/calendar')
        response['Content-Disposition'] = f'attachment; filename="{self.timetable.name}.ics"'
        return response
    
    def _generate_ics_content(self) -> str:
        """Generate ICS calendar content"""
        lines = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//AI Academic Timetable Scheduler//EN",
            f"X-WR-CALNAME:{self.timetable.name}",
            "X-WR-TIMEZONE:UTC",
        ]
        
        # Calculate start date of academic year (assuming it starts in August)
        year = int(self.timetable.academic_year.split('-')[0])
        start_date = datetime(year, 8, 1)  # Adjust as needed
        
        # Generate events for each session
        for session in self.sessions:
            # Calculate the actual date for this session
            days_to_add = session.day_of_week
            session_date = start_date + timedelta(days=days_to_add)
            
            # Create datetime objects
            start_datetime = datetime.combine(session_date.date(), session.start_time)
            end_datetime = datetime.combine(session_date.date(), session.end_time)
            
            # Format for ICS
            start_str = start_datetime.strftime('%Y%m%dT%H%M%S')
            end_str = end_datetime.strftime('%Y%m%dT%H%M%S')
            
            # Create event
            summary = f"{session.subject.code} - {session.class_group}"
            description = f"Subject: {session.subject.name}\\n"
            if session.teacher:
                description += f"Teacher: {session.teacher.user.get_full_name()}\\n"
            if session.room:
                description += f"Room: {session.room.name}\\n"
            
            lines.extend([
                "BEGIN:VEVENT",
                f"DTSTART:{start_str}",
                f"DTEND:{end_str}",
                f"SUMMARY:{summary}",
                f"DESCRIPTION:{description}",
                f"LOCATION:{session.room.name if session.room else ''}",
                f"UID:{session.id}@timetable-scheduler.com",
                "END:VEVENT"
            ])
        
        lines.append("END:VCALENDAR")
        return '\n'.join(lines)
    
    def _get_time_slots(self) -> List[Dict]:
        """Get all time slots for the institution"""
        time_slots = []
        start_time = datetime.combine(datetime.today(), self.timetable.institution.start_time)
        end_time = datetime.combine(datetime.today(), self.timetable.institution.end_time)
        lunch_start = datetime.combine(datetime.today(), self.timetable.institution.lunch_break_start)
        lunch_end = datetime.combine(datetime.today(), self.timetable.institution.lunch_break_end)
        
        slot_duration = timedelta(minutes=self.timetable.institution.slot_duration)
        current_time = start_time
        
        while current_time + slot_duration <= end_time:
            slot_end = current_time + slot_duration
            
            # Skip lunch break
            if not (current_time >= lunch_start and slot_end <= lunch_end):
                time_slots.append({
                    'start_time': current_time.time(),
                    'end_time': slot_end.time()
                })
            
            current_time = slot_end
        
        return time_slots
