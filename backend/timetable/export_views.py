"""
Export views for timetable data in various formats
"""

import io
import json
from datetime import datetime, timedelta
from django.http import HttpResponse, JsonResponse
from django.template.loader import render_to_string
from django.conf import settings
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status

# Import libraries for different export formats
try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from PIL import Image, ImageDraw, ImageFont
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

from .models import Timetable, TimetableSession, Institution
from .serializers import TimetableSessionSerializer


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_timetable_pdf(request, timetable_id):
    """Export timetable as PDF"""
    if not REPORTLAB_AVAILABLE:
        return JsonResponse({
            'error': 'PDF export not available. Install reportlab: pip install reportlab'
        }, status=500)
    
    try:
        timetable = Timetable.objects.get(id=timetable_id)
        sessions = TimetableSession.objects.filter(timetable=timetable).select_related(
            'subject', 'teacher__user', 'room', 'class_group'
        ).order_by('day_of_week', 'start_time')
        
        # Create PDF
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center
        )
        
        # Title
        title = Paragraph(f"Timetable: {timetable.name}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 12))
        
        # Group sessions by day
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        
        for day_num in range(1, 8):  # 1=Monday, 7=Sunday
            day_sessions = sessions.filter(day_of_week=day_num)
            if not day_sessions.exists():
                continue
            
            # Day header
            day_title = Paragraph(f"<b>{days[day_num-1]}</b>", styles['Heading2'])
            elements.append(day_title)
            elements.append(Spacer(1, 6))
            
            # Create table data
            table_data = [
                ['Time', 'Subject', 'Teacher', 'Room', 'Class', 'Type']
            ]
            
            for session in day_sessions:
                table_data.append([
                    f"{session.start_time.strftime('%H:%M')}-{session.end_time.strftime('%H:%M')}",
                    f"{session.subject.code}\n{session.subject.name}",
                    f"{session.teacher.user.first_name} {session.teacher.user.last_name}",
                    session.room.code,
                    session.class_group.name,
                    session.get_session_type_display()
                ])
            
            # Create table
            table = Table(table_data, colWidths=[1*inch, 2*inch, 1.5*inch, 0.8*inch, 0.8*inch, 1*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        
        # Return response
        response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}.pdf"'
        return response
        
    except Timetable.DoesNotExist:
        return JsonResponse({'error': 'Timetable not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_timetable_excel(request, timetable_id):
    """Export timetable as Excel"""
    if not OPENPYXL_AVAILABLE:
        return JsonResponse({
            'error': 'Excel export not available. Install openpyxl: pip install openpyxl'
        }, status=500)
    
    try:
        timetable = Timetable.objects.get(id=timetable_id)
        sessions = TimetableSession.objects.filter(timetable=timetable).select_related(
            'subject', 'teacher__user', 'room', 'class_group'
        ).order_by('day_of_week', 'start_time')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Timetable"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws.merge_cells('A1:G1')
        ws['A1'] = f"Timetable: {timetable.name}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Day', 'Time', 'Subject Code', 'Subject Name', 'Teacher', 'Room', 'Class', 'Type']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        row = 4
        
        for session in sessions:
            ws.cell(row=row, column=1, value=days[session.day_of_week-1]).border = border
            ws.cell(row=row, column=2, value=f"{session.start_time.strftime('%H:%M')}-{session.end_time.strftime('%H:%M')}").border = border
            ws.cell(row=row, column=3, value=session.subject.code).border = border
            ws.cell(row=row, column=4, value=session.subject.name).border = border
            ws.cell(row=row, column=5, value=f"{session.teacher.user.first_name} {session.teacher.user.last_name}").border = border
            ws.cell(row=row, column=6, value=session.room.code).border = border
            ws.cell(row=row, column=7, value=session.class_group.name).border = border
            ws.cell(row=row, column=8, value=session.get_session_type_display()).border = border
            row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        # Return response
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}.xlsx"'
        return response
        
    except Timetable.DoesNotExist:
        return JsonResponse({'error': 'Timetable not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_timetable_png(request, timetable_id):
    """Export timetable as PNG image"""
    if not PIL_AVAILABLE:
        return JsonResponse({
            'error': 'PNG export not available. Install Pillow: pip install Pillow'
        }, status=500)
    
    try:
        timetable = Timetable.objects.get(id=timetable_id)
        sessions = TimetableSession.objects.filter(timetable=timetable).select_related(
            'subject', 'teacher__user', 'room', 'class_group'
        ).order_by('day_of_week', 'start_time')
        
        # Image dimensions
        width, height = 1200, 800
        img = Image.new('RGB', (width, height), color='white')
        draw = ImageDraw.Draw(img)
        
        # Try to load a font
        try:
            title_font = ImageFont.truetype("arial.ttf", 24)
            header_font = ImageFont.truetype("arial.ttf", 16)
            text_font = ImageFont.truetype("arial.ttf", 12)
        except:
            # Fallback to default font
            title_font = ImageFont.load_default()
            header_font = ImageFont.load_default()
            text_font = ImageFont.load_default()
        
        # Title
        title_text = f"Timetable: {timetable.name}"
        title_bbox = draw.textbbox((0, 0), title_text, font=title_font)
        title_width = title_bbox[2] - title_bbox[0]
        draw.text(((width - title_width) // 2, 20), title_text, fill='black', font=title_font)
        
        # Create a simple grid layout
        start_y = 80
        row_height = 25
        col_widths = [100, 120, 150, 200, 150, 80, 100, 80]
        
        # Headers
        headers = ['Day', 'Time', 'Subject Code', 'Subject Name', 'Teacher', 'Room', 'Class', 'Type']
        x = 20
        for i, header in enumerate(headers):
            draw.rectangle([x, start_y, x + col_widths[i], start_y + row_height], outline='black', fill='lightgray')
            draw.text((x + 5, start_y + 5), header, fill='black', font=header_font)
            x += col_widths[i]
        
        # Data rows
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']
        y = start_y + row_height
        
        for session in sessions:
            if y > height - 50:  # Stop if we're running out of space
                break
                
            x = 20
            row_data = [
                days[session.day_of_week-1],
                f"{session.start_time.strftime('%H:%M')}-{session.end_time.strftime('%H:%M')}",
                session.subject.code,
                session.subject.name[:20] + "..." if len(session.subject.name) > 20 else session.subject.name,
                f"{session.teacher.user.first_name} {session.teacher.user.last_name}"[:15],
                session.room.code,
                session.class_group.name,
                session.get_session_type_display()
            ]
            
            for i, data in enumerate(row_data):
                draw.rectangle([x, y, x + col_widths[i], y + row_height], outline='black', fill='white')
                draw.text((x + 5, y + 5), str(data), fill='black', font=text_font)
                x += col_widths[i]
            
            y += row_height
        
        # Save to buffer
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        
        # Return response
        response = HttpResponse(buffer.getvalue(), content_type='image/png')
        response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}.png"'
        return response
        
    except Timetable.DoesNotExist:
        return JsonResponse({'error': 'Timetable not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def export_timetable_ics(request, timetable_id):
    """Export timetable as ICS calendar file"""
    try:
        timetable = Timetable.objects.get(id=timetable_id)
        sessions = TimetableSession.objects.filter(timetable=timetable).select_related(
            'subject', 'teacher__user', 'room', 'class_group'
        )
        
        # Create ICS content
        ics_content = [
            "BEGIN:VCALENDAR",
            "VERSION:2.0",
            "PRODID:-//Demo Timetable//Timetable Export//EN",
            f"X-WR-CALNAME:{timetable.name}",
            "X-WR-TIMEZONE:UTC",
        ]
        
        # Add events for each session
        for session in sessions:
            # Calculate dates for the semester (assuming weekly recurring)
            start_date = timetable.start_date
            
            # Find the first occurrence of this day of week
            days_ahead = session.day_of_week - start_date.weekday() - 1
            if days_ahead < 0:
                days_ahead += 7
            first_occurrence = start_date + timedelta(days=days_ahead)
            
            # Create event
            event_start = datetime.combine(first_occurrence, session.start_time)
            event_end = datetime.combine(first_occurrence, session.end_time)
            
            ics_content.extend([
                "BEGIN:VEVENT",
                f"UID:{session.id}@timetable.demo",
                f"DTSTART:{event_start.strftime('%Y%m%dT%H%M%S')}",
                f"DTEND:{event_end.strftime('%Y%m%dT%H%M%S')}",
                f"SUMMARY:{session.subject.code} - {session.subject.name}",
                f"DESCRIPTION:Teacher: {session.teacher.user.first_name} {session.teacher.user.last_name}\\nRoom: {session.room.code}\\nClass: {session.class_group.name}\\nType: {session.get_session_type_display()}",
                f"LOCATION:{session.room.code} - {session.room.name}",
                f"RRULE:FREQ=WEEKLY;UNTIL={timetable.end_date.strftime('%Y%m%d')}T235959",
                "END:VEVENT",
            ])
        
        ics_content.append("END:VCALENDAR")
        
        # Return response
        response = HttpResponse('\n'.join(ics_content), content_type='text/calendar')
        response['Content-Disposition'] = f'attachment; filename="timetable_{timetable.id}.ics"'
        return response
        
    except Timetable.DoesNotExist:
        return JsonResponse({'error': 'Timetable not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
